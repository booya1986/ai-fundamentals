"""
Generate course-improvement-report.xlsx from the AI Fundamentals course improvement report.
Produces a multi-sheet, formatted Excel workbook.
"""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ── Palette ──────────────────────────────────────────────────────────────────
RED_DARK   = "B31B1B"   # bank red (header text on dark bg)
RED_FILL   = "C0392B"   # header background
RED_LIGHT  = "FDECEA"   # priority High row tint
AMB_LIGHT  = "FFF3E0"   # priority Medium row tint
GRN_LIGHT  = "E8F5E9"   # priority Low row tint
GREY_HDR   = "2C2C2C"   # dark grey for sheet headers
WHITE      = "FFFFFF"
LIGHT_GREY = "F5F5F5"
MID_GREY   = "E0E0E0"
ACCENT_BLUE= "1565C0"   # sprint badge colour

def header_font(bold=True, size=11, color=WHITE):
    return Font(bold=bold, size=size, color=color, name="Calibri")

def body_font(bold=False, size=10, color="000000"):
    return Font(bold=bold, size=size, color=color, name="Calibri")

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def border(style="thin"):
    s = Side(style=style, color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def wrap(horizontal="left", vertical="top", wrap=True):
    return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap)

def set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

def apply_row_style(ws, row_num, bg_hex, bold=False):
    for cell in ws[row_num]:
        if cell.value is not None or True:
            cell.fill = fill(bg_hex)
            cell.font = body_font(bold=bold)
            cell.alignment = wrap()
            cell.border = border()

def write_header_row(ws, row_num, headers, bg=RED_FILL, fg=WHITE, size=10):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col, value=h)
        cell.fill = fill(bg)
        cell.font = Font(bold=True, size=size, color=fg, name="Calibri")
        cell.alignment = wrap(horizontal="center", vertical="center")
        cell.border = border()

def priority_fill(p):
    if p and "High" in p:
        return RED_LIGHT
    if p and "Medium" in p:
        return AMB_LIGHT
    return GRN_LIGHT

# ─────────────────────────────────────────────────────────────────────────────
#  DATA
# ─────────────────────────────────────────────────────────────────────────────

# Sheet 1 — Master Recommendations Table
# Columns: ID | Category | Module | Current State | Issue / Gap | Recommendation |
#          New Format | Priority | Effort | Impact | Sprint | Status

RECOMMENDATIONS = [
    # ── DIAGNOSTIC ──────────────────────────────────────────────────────────
    ("R-01", "Diagnostic & Routing", "M0",
     "Single-dimension diagnostic quiz; no adaptive routing",
     "All employees proceed through identical content regardless of prior AI experience. Over-qualified learners disengage; under-qualified learners are overwhelmed later.",
     "Redesign to a 3-dimension, 10-item pre-assessment measuring AI Knowledge, Tool Proficiency, and Security Awareness. Route learners to Foundation / Standard / Accelerated tracks. Display personalised profile to each learner.",
     "Interactive quiz (in-app)", "High", "Medium", "High", "Sprint 1"),

    # ── M0 ──────────────────────────────────────────────────────────────────
    ("R-02", "Content & Format", "M0",
     "Intro video frames course as 'a journey into AI' for beginners",
     "Many employees already use Copilot Chat daily. The opening framing signals the course is not relevant to them.",
     "Rewrite intro video to open with the employee's existing reality: 'You already have AI in your work environment. This course makes sure you use it smartly, safely, and effectively.' Acknowledge prior use from the first sentence.",
     "Video (rewrite script)", "High", "Low", "High", "Sprint 1"),

    ("R-03", "Content & Format", "M0",
     "No course orientation visual",
     "Learners start without understanding the shape of the journey or which modules connect to which goals.",
     "Add a 60-second animated course map showing all 9 modules, which connect to Copilot use, which introduces additional tools, which covers safety, and where the capstone sits.",
     "Short animated video (60 sec)", "Medium", "Low", "Medium", "Sprint 3"),

    # ── M1 ──────────────────────────────────────────────────────────────────
    ("R-04", "Content & Format", "M1 / l1",
     "Standalone reading lesson (4 min, 3 keypoints) before WordPredictionSim",
     "The reading explains word-by-word prediction and statistical engines — exactly what the WordPredictionSim demonstrates experientially two minutes later. The lesson duplicates what the simulation teaches, adding time without adding learning.",
     "Remove M1/l1 as a standalone lesson. Merge its core concept (3 sentences) into the WordPredictionSim intro screen. Net saving: 4 minutes and one navigation step.",
     "Merge into existing sim (no new production)", "High", "Low", "High", "Sprint 1"),

    ("R-05", "Content & Format", "M1 / lmap",
     "Dense reading (6 min, 4 very long keypoints) covering AI hierarchy, knowledge cutoff, Attention mechanism, and tokenisation",
     "Four abstract, interconnected concepts on a mobile screen. Each keypoint is 100–200 words. The Attention mechanism section alone describes a mathematical process. Mobile learners skim or abandon this lesson.",
     "Replace with 2 short animated explainer videos (3 min each). Video A: The AI Map — zooming animation through AI > ML > DL > GenAI > LLM, ending with the knowledge cutoff visualised as a stopped calendar. Video B: Tokens and Context — text breaking into token blocks, Attention weights visualised, ending with the connection to prompt quality.",
     "2 × animated video (3 min each)", "High", "Medium", "High", "Sprint 3"),

    ("R-06", "Content & Format", "M1 / lmap",
     "No audio version of foundational AI concepts",
     "Mobile learners on commutes cannot engage with reading or video equally. An audio version of the conceptual content extends reach without additional lesson time.",
     "Generate a 5-minute Audio Overview in NotebookLM from the M1/lmap transcript. Embed as an optional 'Listen first' option at the top of the module. Simultaneously models the NotebookLM tool for employees.",
     "Audio Overview (NotebookLM — low effort)", "Medium", "Low", "Medium", "Sprint 3"),

    ("R-07", "New Interaction", "M1",
     "No trust calibration practice after conceptual lessons",
     "The module teaches WHY AI hallucinations happen but not how to DETECT them in professional outputs. Employees finish M1 with the concept but without the judgment skill.",
     "Add 'Spot the Problem' — 5 Copilot-style output cards showing banking-context responses. Learner decides 'Use as-is' or 'Verify first' for each. After each card: one-sentence explanation of the reasoning. This trains calibrated trust through practice.",
     "5-card interactive (new)", "High", "Medium", "High", "Sprint 2"),

    ("R-08", "Content & Format", "M1 / l4 (ToolsDifferences)",
     "ToolsDifferences simulation does not include JARVIS",
     "JARVIS is introduced only in M2, but employees will use Copilot in an exercise immediately after this sim. They need to know that for sensitive tasks there is an internal alternative.",
     "Add JARVIS as a fifth tool in the ToolsDifferences simulation, with 'data stays inside the bank' as its defining characteristic. One extra column/card in the existing sim.",
     "Sim update (extend existing)", "Medium", "Low", "Medium", "Sprint 2"),

    # ── M2 ──────────────────────────────────────────────────────────────────
    ("R-09", "Content & Format", "M2 / l1",
     "Reading (6 min, 5 keypoints) describing each tool in paragraph form",
     "Teaches tool selection as a parallel choice between 5 equal tools. The real decision is sequential (starting from Copilot as default) and gate-keeps on data sensitivity first. Also conflates JARVIS the platform with the Compliance Agent.",
     "Replace reading with an interactive decision flow: a clickable branch that walks through one real scenario, asking data sensitivity, document volume, and output type questions to arrive at the right tool. Follow with a tappable reference card grid — one card per tool, permanently accessible as a job aid.",
     "Interactive decision tree + reference card grid", "High", "Medium", "High", "Sprint 2"),

    ("R-10", "Content Accuracy", "M2 / l1",
     "JARVIS described as a policy lookup tool ('offline, שאלות נהלים, מפנה ל-POINT')",
     "This describes the Compliance Agent — one application built on the JARVIS platform. Employees who only know JARVIS as a policy tool will not use it for general tasks involving sensitive data, which is its primary value.",
     "Rewrite JARVIS description: 'JARVIS is the bank's internal LLM platform — same general capabilities as Copilot Chat, but data never leaves the bank. The bank is also building specialised agents on JARVIS, including a Compliance Agent trained on all bank compliance documents. Use JARVIS for any task where data is sensitive.'",
     "Text rewrite (no production)", "High", "Low", "High", "Sprint 1"),

    ("R-11", "Assessment", "M2 / quiz",
     "Q2 asks which tool for drafting a client letter; correct answer is Copilot",
     "The scenario implies the letter is generic. In reality, if the letter contains the client's name and account number, JARVIS is the correct tool. The question teaches the wrong decision rule.",
     "Rewrite Q2 to include 'the letter contains the client's full name and account number.' Correct answer becomes JARVIS. Explanation acknowledges that anonymise-then-Copilot is also valid, but JARVIS is safest.",
     "Quiz item rewrite", "High", "Low", "High", "Sprint 1"),

    ("R-12", "Assessment", "M2 / quiz",
     "Q4 matching: JARVIS description reflects only the Compliance Agent",
     "The matching pair for JARVIS says 'פנימי, offline, שאלות נהלים ומפנה ל-POINT' — which is the Compliance Agent, not the platform.",
     "Update JARVIS matching description to: 'פלטפורמת LLM פנימית — כל יכולות Copilot, הנתונים לא יוצאים. מארחת סוכנים כמו סוכן הציות.'",
     "Quiz item update", "High", "Low", "High", "Sprint 1"),

    ("R-13", "Assessment", "M2 / quiz",
     "Q5 fill-blank for JARVIS uses 5-word bank with 2 obvious wrong answers (Copilot, POINT)",
     "Effectively a 3-option recognition task. No recall required.",
     "Replace with a scenario application: employee asks JARVIS about current interest rates and gets an answer from a 2023 policy. What should they know and do? Tests understanding of RAG limitations, not label recognition.",
     "Quiz item rewrite", "Medium", "Low", "Medium", "Sprint 1"),

    # ── M3 ──────────────────────────────────────────────────────────────────
    ("R-14", "Content & Format", "M3 / l2",
     "Reading (8 min) explaining the RICE framework declaratively",
     "RICE is a construction framework — teaching it through reading means learners observe instead of practise. The subsequent Copilot exercises assume RICE competency but do not build it incrementally.",
     "Replace reading with a RICE Builder interactive: a real banking task shown, four input fields (R, I, C, E) with guidance, a live preview of the assembled prompt, and a 'Compare' button showing output vs. a weak one-sentence prompt. Keep downloadable RICE Quick Card as a reference text.",
     "RICE Builder interactive (4 fields + live preview)", "Medium", "Medium", "High", "Sprint 2"),

    ("R-15", "Content & Format", "M3",
     "No audio reinforcement for highest-value module",
     "RICE is the most directly applicable skill in the course. Commute-friendly audio reinforcement would extend retention.",
     "Generate a 5-minute NotebookLM Audio Overview from the RICE module content and banking examples. Optional listen at M3 entry or post-exercise.",
     "Audio Overview (NotebookLM — low effort)", "Medium", "Low", "Medium", "Sprint 3"),

    ("R-16", "Content & Format", "M3 / l3",
     "Video 'Weak vs. Good Prompt' ends without a practice bridge",
     "The video demonstrates the difference; learners move directly to a live Copilot exercise. A one-question bridge would ensure the concept is active before tool use.",
     "After the video, add one MCQ: 'This is a weak prompt for a banking task — which RICE element is missing?' Immediate feedback, then proceed to exercise.",
     "Single bridge question (new)", "Medium", "Low", "Medium", "Sprint 1"),

    ("R-17", "Assessment", "M3 / quiz",
     "No application-level item — all items test recall of RICE labels",
     "A learner who scores 100% on the quiz may still write a weak prompt. No item tests whether they can evaluate or apply the framework.",
     "Add scenario item: employee sent Copilot a vague request and got a generic result. Learner identifies which RICE element is most clearly missing. Four options with plausible distractors.",
     "New scenario quiz item", "High", "Low", "High", "Sprint 1"),

    ("R-18", "Content & Format", "M3 / exercises",
     "Reflection fields in both Copilot exercises are generic ('what worked?')",
     "Generic reflection fields are easy to complete with minimal thought. They do not require the learner to engage with the RICE framework.",
     "Replace reflection fields with targeted questions: Exercise l4: 'Which RICE element made the biggest difference — and what would have been missing without it?' Exercise l5: 'Compare your output to what a one-sentence prompt would have produced. What specifically changed?'",
     "Text update (no production)", "Medium", "Low", "Medium", "Sprint 1"),

    # ── M4 ──────────────────────────────────────────────────────────────────
    ("R-19", "Content & Format", "M4 / lcap",
     "Reading (5 min) 'Capabilities vs. Limitations: when to trust and when not to'",
     "Trust calibration is a judgment skill. A declarative list of when to trust AI teaches the rule but not the judgment. Employees who read this will still accept a hallucinated interest rate at their desk.",
     "Replace with 'Trust Meter' — 5 Copilot-style output cards for banking-context requests. Learner taps 'Use as-is' or 'Verify first' for each. After each: one-sentence explanation. Cards cover: safe draft email, quoted interest rate, regulation with a date, client recommendation, reworded announcement.",
     "5-card Trust Meter interaction (new)", "High", "Medium", "High", "Sprint 2"),

    ("R-20", "Content & Format", "M4 / videos",
     "No chapter navigation on videos in longest module (35 min total)",
     "Two videos in the second-longest module are the primary mobile dropout points. Without chapter markers, interrupted learners must restart from the beginning.",
     "Add WebVTT chapter tracks to both M4 videos, with chapter breaks every 2–3 minutes. Display chapter list as a collapsible menu above each video.",
     "Chapter navigation (WebVTT — technical update)", "Medium", "Low", "Medium", "Sprint 4"),

    ("R-21", "Content & Format", "M4 / exercise",
     "Copilot exercise does not connect to the employee's existing Copilot use",
     "The exercise treats Copilot as a new tool to learn. Employees use it daily. Connecting the exercise to prior experience would improve transfer.",
     "Add opening frame: 'Think of a Copilot task you did this week. Keep it in mind.' Add reflection field: 'How does this exercise compare to how you usually use Copilot? What would you do differently for that task?'",
     "Text update (no production)", "Medium", "Low", "Medium", "Sprint 1"),

    ("R-22", "Assessment", "M4 / quiz",
     "No trust calibration scenario item",
     "The module teaches when to trust AI output. The quiz does not test whether the learner can apply this judgment to a real situation.",
     "Add scenario: Copilot confidently quotes the Bank of Israel's current prime lending rate. Learner is about to include this in a client document. What should they do? Four options including 'ask Copilot to confirm' (wrong) as a plausible distractor.",
     "New scenario quiz item", "High", "Low", "High", "Sprint 1"),

    # ── M5 ──────────────────────────────────────────────────────────────────
    ("R-23", "Content & Format", "M5",
     "Module opens directly with 2 Copilot exercises — no conceptual content",
     "M5 is the only module that violates the Tell-Show-Do sequence used consistently elsewhere. Employees encounter two unfamiliar tools with no framing. Exercises produce frustration rather than competency.",
     "Add before exercises: (1) NotebookLM screen-capture walkthrough video (4 min) — upload a fictional regulatory document, ask a question, see cited answer, generate Audio Overview. (2) Napkin screen-capture (90 sec) — paste a loan process description, watch a diagram appear.",
     "2 screen-capture videos (4 min + 90 sec)", "High", "Medium", "High", "Sprint 3"),

    ("R-24", "Content & Format", "M5",
     "No explanation of why these tools exist when employees already have Copilot Chat",
     "Employees with Copilot Chat as their daily tool need a compelling answer to 'why would I use something else?' before they'll engage with new tools.",
     "Add a short framing section: Copilot Chat handles up to 3 files / 8,000 characters. NotebookLM handles many more long documents. Napkin produces diagrams that Copilot cannot. These extend Copilot; they do not replace it.",
     "Short text section (no production)", "High", "Low", "High", "Sprint 1"),

    ("R-25", "Content & Format", "M5",
     "No audio version of M5 content",
     "A module about NotebookLM that includes a NotebookLM Audio Overview of its own content is pedagogically elegant — it teaches the content and demonstrates the tool simultaneously.",
     "Generate a 3-minute Audio Overview from the M5 explainer content. Embed as optional 'Listen first' at module entry.",
     "Audio Overview (NotebookLM — low effort)", "Medium", "Low", "Medium", "Sprint 3"),

    ("R-26", "Assessment", "M5 / quiz",
     "Quiz tests recall without conceptual foundation",
     "With no prior instruction, the M5 quiz can only test what learners discovered by doing the exercises. It functions as exposure, not measurement.",
     "After adding explainer videos (R-23), rewrite quiz to include a scenario: an employee has 15 regulatory documents to synthesise before a meeting and wants to listen to a summary on her commute. Which tool? Four options with plausible distractors.",
     "New scenario quiz item (after R-23 is implemented)", "Medium", "Low", "Medium", "Sprint 3"),

    # ── M6 ──────────────────────────────────────────────────────────────────
    ("R-27", "Content & Format", "M6 / video",
     "Ethics video is purely preventive — does not address employees who may already be sharing sensitive data with Copilot",
     "Since Copilot Chat is the daily tool, some employees may already have shared confidential information with it before taking this course. Treating M6 as purely preventive misses this population.",
     "Add a 60-second segment: 'If you've already used Copilot Chat with real customer information — you're not alone. Here is what to do going forward.' Acknowledge without shaming. Provide a clear path: these tasks move to JARVIS from now on.",
     "Video segment addition (60 sec)", "High", "Low", "High", "Sprint 3"),

    ("R-28", "Content & Format", "M6 / EthicsCyber sim",
     "Simulation consequences are generic ('this could cause a problem')",
     "Generic consequences are less motivating than specific, credible institutional ones. 'Something bad could happen' lands differently from 'this would require a mandatory incident report.'",
     "Replace generic consequence text with specific Bank Hapoalim policy references: 'This action could trigger a mandatory incident report under Bank Information Security Policy [number].' Coordinate with Information Security team for accuracy.",
     "Sim text update", "High", "Low", "High", "Sprint 1"),

    ("R-29", "New Interaction", "M6",
     "No practical data classification tool after the simulation",
     "The simulation teaches ethical decision-making. Employees also need a moment-of-use classification tool — 'is this data I'm about to paste into Copilot safe?'",
     "Add 'Data Classifier' micro-tool: employee describes their task/data in plain language; tool classifies as Green (safe for external AI) / Yellow (anonymise first) / Red (JARVIS only). Keyword-pattern classification. Not a quiz — a practice tool.",
     "Data Classifier micro-tool (new)", "High", "Medium", "High", "Sprint 2"),

    # ── M7 ──────────────────────────────────────────────────────────────────
    ("R-30", "Content & Format", "M7",
     "No connection between Custom Instructions and the bank's own AI agent work",
     "The bank is building specialised agents on JARVIS (e.g. Compliance Agent). This gives the 'architect' framing real institutional grounding that is currently missing.",
     "Add a 2-minute section: 'The bank's tech team is building agents on JARVIS using the same principles you're learning — define the role, the knowledge, the rules. Your Custom Instructions are the personal version of that.' Connect personal skill to organisational context.",
     "Short video or reading (2 min)", "Medium", "Low", "High", "Sprint 3"),

    ("R-31", "New Interaction", "M7",
     "No structured practice before the live Copilot exercise for Custom Instructions",
     "The Copilot exercise assumes employees can construct Custom Instructions. There is no intermediate step where they practise the structure before working in the real tool.",
     "Add Custom Instructions Builder interactive (3 fields): What role should Copilot play? What should it always do? What should it never do? Live preview assembles the instructions. One-click copy for use in real Copilot.",
     "Custom Instructions Builder (3-field + preview)", "Medium", "Medium", "Medium", "Sprint 2"),

    ("R-32", "Assessment", "M7 / quiz",
     "No application-level item — quiz tests recall of what Custom Instructions are",
     "Employees need to demonstrate they can construct instructions, not just describe them.",
     "Add an open-response item: employee wants Copilot to always respond in formal Hebrew, use bullet points for summaries, and never include client names. Learner writes the Custom Instruction draft. Model answer shown for self-assessment.",
     "New open-response quiz item", "Medium", "Low", "Medium", "Sprint 1"),

    # ── M8 ──────────────────────────────────────────────────────────────────
    ("R-33", "Content & Format", "M8",
     "Capstone depends entirely on Bino Custom GPT — pending IT rollout",
     "The course's highest-cognitive-demand activity is inaccessible to employees without Bino access. No fallback exists. This is a single point of failure for the capstone.",
     "Build a self-guided fallback: same 6 fields, plus a collapsible 'Bino would ask you' section under each field with 2 Socratic follow-up questions. Downloadable PDF version available. Flag completions via this path for L&D follow-up.",
     "Fallback interactive + PDF (new)", "High", "Low", "High", "Sprint 2"),

    ("R-34", "Content & Format", "M8",
     "Capstone is a private, isolated exercise with no social dimension",
     "Employees have no way to share their capstone plan with their manager or peers. Transfer to on-the-job behaviour requires environmental support — managers who know what was learned and can reinforce it.",
     "Add share mechanism at M8 exit: pre-filled Teams message or email with a summary of the learner's capstone plan (fields 1–2) and a prompt to discuss implementation. One-click send.",
     "Share mechanism (Teams/email integration)", "High", "Medium", "High", "Sprint 2"),

    ("R-35", "Content & Format", "M8",
     "No social proof or peer context before the capstone",
     "Learners enter the capstone without seeing how real colleagues have applied course content. Peer examples at this moment increase completion and ambition of the final plan.",
     "Add 3 × 90-second peer story videos at M8 entry. Real employees (named + role) describe one specific change they made using course content — concrete and replicable ('I'm a mortgage advisor; I use this RICE prompt every time a client asks about refinancing. Here's the exact structure.').",
     "3 × peer story videos (90 sec each, phone recording)", "Medium", "Low", "High", "Sprint 3"),

    # ── CROSS-CUTTING: Assessment quality ────────────────────────────────────
    ("R-36", "Assessment", "M1 / quiz",
     "Q4: Why do ChatGPT and Copilot give different answers? — all 3 wrong options immediately dismissible",
     "Distractors: 'websites designed in different colours,' 'internet is faster on one tool,' 'no real difference.' A learner with any Copilot experience eliminates all three instantly. Zero discriminating power.",
     "Replace with plausible misconceptions: (a) 'Copilot is always connected to real-time internet so it has more current info' (partially true in some modes); (b) 'they use the same knowledge base but different writing styles' (common misconception); (c) 'Claude was trained on books, Copilot on the internet, making Claude more accurate for formal content' (sounds authoritative, factually oversimplified).",
     "Quiz item rewrite", "High", "Low", "High", "Sprint 1"),

    ("R-37", "Assessment", "M1 / quiz",
     "Q3: Tokenisation fill-blank — 4-word bank, 2 obviously wrong options; recognition not recall",
     "With 'pixels' and 'bytes' as distractors, this is effectively a 2-option recognition task. Anyone who glanced at the lesson can answer correctly.",
     "Replace with a free-text production item: 'A colleague says the AI read and understood my document. In one sentence, explain why this is misleading using what you know about tokens and probability.' Show a model answer after submission for self-evaluation.",
     "Quiz item format change (free text)", "High", "Low", "High", "Sprint 1"),

    ("R-38", "Assessment", "M1 / quiz",
     "Q2: T/F 'LLMs understand meaning like humans' — binary format hides level of comprehension",
     "A learner who answers False might mean 'no understanding at all' (correct) or 'it understands differently' (imprecise). Same answer, different comprehension levels.",
     "Replace with scenario MCQ: 'A colleague says Copilot understood my report exactly like I would. Which is most accurate?' Four options ranging from mechanistic accuracy (correct) to anthropomorphism (wrong) to denial of Hebrew capability (wrong).",
     "Quiz item format change (MCQ)", "Medium", "Low", "Medium", "Sprint 1"),

    # ── CROSS-CUTTING: Missing content ───────────────────────────────────────
    ("R-39", "Missing Content", "M1 + M4",
     "No hallucination detection practice anywhere in the course",
     "The course explains why hallucinations happen (M1 WordPredictionSim) but never teaches employees to detect them in actual AI outputs. This is the most consequential missing applied skill.",
     "Add 'Fact-Check This' interaction in M1 (after trust calibration cards): 5 AI-generated banking-context outputs, each with one plausible but verifiable error. Learners identify the suspicious claim. After each: the error revealed with an explanation of why the model produced it.",
     "5-item hallucination detection interactive (new)", "High", "Medium", "High", "Sprint 2"),

    ("R-40", "Missing Content", "M3 + M4",
     "No prompt failure diagnosis — course teaches how to write a good prompt but not what to do when one fails",
     "Prompt failures are common. Employees who have no framework for diagnosing a bad output will either give up on AI or accept a poor result.",
     "Add 'What Went Wrong?' micro-lesson in M4: 3 examples of real Copilot failures. For each: the weak prompt, the bad output, and the task — identify which RICE element failed and how to fix the prompt. Turns RICE from a construction tool into a diagnostic tool.",
     "3-example diagnostic interaction (new)", "High", "Medium", "High", "Sprint 2"),

    ("R-41", "Missing Content", "M4",
     "No AI workflow design content — course teaches individual interactions, not multi-step task integration",
     "Productivity gains from AI come from integrating it into workflows, not from individual interactions. This skill is currently untaught.",
     "Add Workflow Mapping exercise in M4: learner describes a recurring multi-step task, then maps each step: AI-assistable? Human-only? Involves sensitive data (→ JARVIS)? Completed template is a personal AI workflow plan — a tangible takeaway.",
     "Workflow Mapping template interaction (new)", "Medium", "Medium", "High", "Sprint 2"),

    ("R-42", "Missing Content", "M4",
     "No 'When Not to Use AI' content — course teaches when AI works, not when it is the wrong tool",
     "Over-reliance is a real risk. Employees need to know that professional judgment includes recognising when human expertise, empathy, or direct conversation is better than AI.",
     "Add a short section in M4 (reading or 3 cards): 3–5 real banking scenarios where AI makes the situation worse — a distressed client call requiring empathy, a complex credit decision with nuanced factors, a regulatory interpretation requiring a qualified human. Short, direct.",
     "3-card or short reading section (new)", "High", "Low", "High", "Sprint 1"),

    # ── CROSS-CUTTING: Transfer and reinforcement ────────────────────────────
    ("R-43", "Post-Course Transfer", "Post-completion",
     "No post-course reinforcement — course ends at M8 with no follow-up",
     "Without retrieval practice, ~70% of course content is inaccessible within a week (Ebbinghaus forgetting curve). M3 (prompting) and M6 (security) are the highest-stakes content for retention.",
     "Configure D+7 and D+30 spaced retrieval quizzes via LMS: 3 items each, completable in under 3 minutes, application-level (not recall). Triggered by automated LMS notification. Log results to learner record.",
     "Spaced retrieval quizzes + LMS automation", "High", "Medium", "High", "Sprint 4"),

    ("R-44", "Post-Course Transfer", "Post-completion",
     "No manager communication kit — managers do not know what employees learned or how to support application",
     "Level 3 behavior transfer almost never occurs without environmental support. When managers do not know what was learned, they cannot reinforce it, prompt application, or remove barriers.",
     "Create a 1-page manager guide: what your employee completed, what they should now be able to do (5 specific behaviours), and 3 suggested conversation prompts/task assignments. Delivered via LMS completion email. PDF and Teams template formats.",
     "Manager kit (PDF + Teams template)", "High", "Low", "High", "Sprint 4"),

    ("R-45", "Post-Course Transfer", "Post-completion",
     "No matched post-test — cannot measure pre/post learning gain",
     "Without a parallel post-test matched to the M0 diagnostic, it is impossible to demonstrate Level 2 learning gain. This is the foundational data point for any ROI or course-continuation decision.",
     "Write 10 parallel post-test items covering the same constructs as the M0 diagnostic but using different surface content. Display at course completion. Configure LMS to report both scores and calculate gain. Show individual gain score on completion certificate.",
     "Post-test (10 items) + LMS reporting config", "High", "Low", "High", "Sprint 4"),

    # ── ACCESSIBILITY ────────────────────────────────────────────────────────
    ("R-46", "Accessibility", "All video lessons",
     "No subtitles or transcripts on any video (7 videos across M1, M3, M4, M6, M7)",
     "Mobile learners in noisy environments, hearing-impaired employees, and non-native Hebrew speakers cannot fully access video content. Subtitles serve all three populations simultaneously.",
     "Generate Hebrew SRT files for all 7 videos using AI transcription (Whisper or equivalent). Review for banking terminology accuracy. Embed as WebVTT subtitle tracks. Add collapsible transcript accordion below each video.",
     "Hebrew subtitles + transcripts (all videos)", "High", "Medium", "High", "Sprint 4"),

    ("R-47", "Accessibility", "All modules",
     "No Arabic language version — some bank employees are Arabic speakers",
     "Arabic-speaking employees who are required to complete this training but lack full Hebrew proficiency will have lower comprehension and lower transfer — not due to lower capability but due to language access.",
     "Translate all text-based content, quiz items, and UI labels into Arabic. Generate Arabic subtitle tracks from translated transcripts. Prioritise M6 (compliance-critical) and M1 (foundational) for first sprint. RTL layout already compatible.",
     "Arabic translation — M6 + M1 first (8 weeks), remaining modules second sprint", "High", "High", "High", "Sprint 4"),

    # ── CERTIFICATION ARCHITECTURE ────────────────────────────────────────────
    ("R-48", "Certification Architecture", "All",
     "No minimum passing standard — course completion equals certification regardless of assessment performance",
     "An employee who clicks through every screen at 2× speed receives the same credential as one who engaged deeply. This is a course completion record, not a competency certification.",
     "Define certification standards: (1) Complete all mandatory modules per learning track, (2) Pass matched post-test at ≥70%, (3) Complete M8 capstone (Bino or self-guided fallback), (4) Maximum 3 post-test attempts before L&D facilitated session. Module quizzes remain learning checkpoints, not certification requirements.",
     "Policy definition + LMS configuration", "High", "Low", "High", "Sprint 0"),

    ("R-49", "Certification Architecture", "All",
     "No certificate artifact with organizational meaning",
     "Completion is recorded in the LMS but produces no artifact that carries meaning outside the LMS. Managers cannot verify certification status. HR cannot record it as a competency.",
     "Design a digital badge compliant with the Open Badges standard: employee name, date, issuer (Bank Hapoalim L&D Academy), competency level (AI Literacy Level 1), skills claimed (prompt engineering, data security, tool selection, responsible AI use), verification link, expiration date (18 months). Badge visible in LMS profile, HR system, and manager dashboard.",
     "Digital badge design + Open Badges integration", "High", "Medium", "High", "Sprint 0"),

    ("R-50", "Certification Architecture", "All",
     "No recertification cycle — certification has no expiry despite rapidly evolving AI tools and policies",
     "AI tools and bank security policies change faster than a static certification can track. A 2026 certification will misrepresent competency by 2028 if never renewed.",
     "Define 18-month certification validity with LMS-triggered recertification notification 2 months before expiry. Recertification path: 20-minute focused update (new tools, policy changes, refreshed post-test) — not the full course. Add trigger-based recertification for major AI platform changes or policy updates, independent of the 18-month cycle.",
     "Recertification policy + LMS automation", "High", "Low", "High", "Sprint 0"),

    ("R-51", "Certification Architecture", "All",
     "No HR system integration — certification exists only in the LMS, invisible to talent and performance systems",
     "If completion is not recorded in the HR system, it cannot be referenced in performance conversations, onboarding requirements, or workforce capability reporting.",
     "Integrate LMS completion data with the bank's HR system to record 'AI Literacy Level 1' as a verified competency on each employee record. Include in new employee onboarding checklist (completion within 30 days). Make visible to HR Business Partners and department heads via aggregated reporting.",
     "LMS–HR system integration (API or data export)", "High", "Medium", "High", "Sprint 0"),

    # ── MANAGER ENABLEMENT ────────────────────────────────────────────────────
    ("R-52", "Manager Enablement", "Pre-launch",
     "Managers receive no briefing before their team is assigned the course",
     "Without a manager preview, managers are as surprised by the certification as their teams. They cannot introduce it, answer questions about it, or position it as meaningful development. Manager disengagement signals to employees that the certification is a compliance exercise.",
     "Create a 2-page manager preview brief: what the course covers, what employees will be able to do, a 30-minute team meeting agenda for introducing the certification, and a plain-language summary of what will change (JARVIS vs. Copilot decisions, RICE prompting, data classification). Distribute 4 weeks before launch.",
     "Manager brief PDF (2-pager + meeting agenda)", "High", "Low", "High", "Sprint 0"),

    ("R-53", "Manager Enablement", "LMS",
     "No manager dashboard — managers cannot track their team's certification progress",
     "Without visibility, managers cannot identify at-risk team members, address non-completion, or have informed activation conversations after their team completes the certification.",
     "Configure manager team dashboard in LMS: team completion rate, module-level progress per employee, post-test scores, pre/post knowledge gain delta, at-risk flags (not started after 2 weeks, or failed post-test twice). Deliver weekly automated progress email to managers during the rollout window.",
     "LMS manager dashboard + automated weekly email", "High", "Medium", "High", "Sprint 0"),

    ("R-54", "Manager Enablement", "Post-completion",
     "No structured activation conversation — managers do not know how to reinforce the certification after completion",
     "The most important manager conversation happens after an employee completes the course. Without a framework, most managers either say nothing or ask a vague 'so how was it?' The certification's Level 3 transfer depends almost entirely on this conversation.",
     "Create a post-certification activation guide: 3 structured questions to ask every team member (what surprised you? what will you try this week? where could AI make a team-level difference?) + 3 suggested follow-up tasks to assign immediately (RICE prompt trial, JARVIS use with client data, workflow mapping task). One-page, practical.",
     "Activation guide (1-page PDF)", "High", "Low", "High", "Sprint 0"),

    ("R-55", "Manager Enablement", "Post-completion",
     "Managers have no guidance on responding to AI data security disclosures from their team",
     "Some team members may disclose (during the activation conversation or otherwise) that they have previously shared sensitive data with external AI tools. Without guidance, managers risk either over-reacting (shame, escalation) or under-reacting (missing a real incident). Neither serves the organization.",
     "Include in the manager brief: how to respond non-punitively when a team member discloses prior sensitive data use with external AI tools (acknowledge, inform about going-forward policy, do not shame); how to model good AI use in team settings; how to establish team norms around when AI is used and when human judgment is explicitly required.",
     "Manager brief section (text + conversation guide)", "High", "Low", "High", "Sprint 0"),

    # ── ORGANIZATIONAL ROLLOUT ────────────────────────────────────────────────
    ("R-56", "Rollout & Change Management", "M0",
     "No executive sponsorship — the certification launches without visible senior leadership context",
     "Mandatory certification without executive sponsorship reads as an HR compliance exercise. When employees see that senior leadership cares enough to explain why, completion quality and engagement improve significantly.",
     "Produce a 3–5 minute video from a named senior executive (CIO, CHRO, or equivalent) explaining why AI literacy is a bank-wide priority now, what every employee will be able to do after completing it, and what it means for customers and colleagues. Embed in M0 as a prominently featured optional resource.",
     "Executive sponsorship video (3–5 min)", "High", "Low", "High", "Sprint 0"),

    ("R-57", "Rollout & Change Management", "Pre-launch",
     "No communication cascade plan — the certification launches without structured employee and manager communication",
     "Without a structured communication plan, employees receive the certification with no context, no advance notice, and no understanding of why it is mandatory. This creates resistance and low-quality completion.",
     "Design and execute a 4-week communication cascade: Week -4: manager brief + dashboard access. Week -2: all-employee announcement (why, how long, deadline, work time allocation confirmed). Week -1: reminder + FAQ + helpdesk contact. Launch day: direct link. Weekly: automated manager progress reports during rollout. Include role-tailored WIIFM messages for 4 employee segments (branch staff, knowledge workers, managers, executives).",
     "Communication cascade plan + FAQ document + 4 role-tailored messages", "High", "Low", "High", "Sprint 0"),

    ("R-58", "Rollout & Change Management", "All",
     "No completion mandate structure — unclear what happens if employees do not complete the certification",
     "Ambiguous accountability signals that the certification is optional in practice. Clear, proportionate, and pre-communicated consequences create the accountability structure a mandatory certification requires.",
     "Define and communicate: 8-week completion window for existing employees, 30 days for new hires, weekly manager progress reports with at-risk flags, week-4 non-completion flag to direct manager, week-8 non-completion escalation to skip-level manager, support (L&D facilitated session) before consequences. Communicate all of this before launch so no employee is surprised.",
     "Mandate policy (internal document) + communication language", "High", "Low", "High", "Sprint 0"),

    # ── POST-CERTIFICATION ECOSYSTEM ─────────────────────────────────────────
    ("R-59", "Post-Certification Ecosystem", "Post-completion",
     "No community or network for employees who want to go further after certification",
     "The certification creates a population of employees who have engaged with AI fundamentals. Without a community, their energy dissipates. Early adopters have nowhere to share, and L&D has no channel for ongoing intelligence about how AI is actually being used in the bank.",
     "At M8 exit, offer an opt-in 'AI Champions' network: a dedicated Teams channel for sharing prompts, use cases, and questions; a monthly AI learning digest from L&D; early access to new approved tools; invitation to contribute peer story videos for future course updates. Champions serve as L&D's advisory group for quarterly content reviews.",
     "AI Champions Teams channel + monthly digest + quarterly check-in", "Medium", "Low", "High", "Sprint 5"),

    ("R-60", "Post-Certification Ecosystem", "Post-completion",
     "No maintained resource hub — job aids and references exist only inside the course and will not be found after completion",
     "Once employees complete the certification, they cannot easily access the Tool Reference Card, RICE Quick Card, or Data Classification Guide when they need them at their desk. Resources buried in a completed LMS module are as good as inaccessible.",
     "Build a Living Resource Hub on the intranet or a pinned Teams channel: Tool Reference Card (updated as approved tools change), RICE Prompt Gallery (seeded with 15–20 role-specific prompts, updated quarterly from Champions contributions), Data Classification Guide (maintained by Information Security), and a 'What's New' feed for tool and policy changes.",
     "Intranet page or pinned Teams channel + quarterly update process", "Medium", "Low", "High", "Sprint 5"),

    # ── GOVERNANCE ───────────────────────────────────────────────────────────
    ("R-61", "Governance", "All",
     "No content governance model — no named owner, no review cadence, no trigger for updates",
     "Without governance, the certification content will be materially outdated within 12 months. AI tools change. Bank policies change. JARVIS capabilities expand. A course that reflects a 2026 reality is a liability by 2027.",
     "Assign a named L&D content owner (not 'L&D' generically). Define review cadences: quarterly review of tool-specific content (M2, M5) and data security content (M6); annual full review; trigger-based update within 4 weeks of any new tool addition, policy change, or major AI capability shift. Implement version control: each update increments the version; employees who completed an earlier version receive a short 'What's New' update (under 10 minutes), not the full course.",
     "Governance policy + content owner assignment + version control process", "High", "Low", "High", "Sprint 5"),

    # ── LEARNING JOURNEY ─────────────────────────────────────────────────────
    ("R-62", "Learning Journey", "M8",
     "Certification ends at M8 with no reference to what comes next — no learning journey context",
     "If this is 'the first competency,' employees who complete it should understand where it sits in a larger framework. Without this, certification feels like a one-time event rather than a foundation for ongoing development.",
     "Add an exit screen at M8 completion: name the certification as 'AI Literacy Level 1' within the bank's AI competency framework. Reference what comes next — Level 2 for advanced users, a Team AI Implementation Guide for managers, a technical track for developers. Even if these do not yet exist, naming them creates an expectation and L&D accountability to deliver. Design a 4-level competency architecture: Level 1 (all employees, current), Level 2 (advanced users), Level 3 (AI team leads), Specialist Track (technical staff).",
     "M8 exit screen + 4-level competency architecture definition", "Medium", "Low", "High", "Sprint 5"),
]


# Sheet 2 — Format Transformation Decision Table
FORMAT_DECISIONS = [
    # (Content, Current Format, Recommended Format, Rationale, Module, Priority)
    ("Intro to AI as statistical engine", "Reading (4 min)", "Merge into WordPredictionSim intro (3 sentences)", "Sim teaches the concept experientially — reading duplicates it", "M1/l1", "High"),
    ("AI hierarchy + knowledge cutoff + Attention + tokens", "Dense reading (6 min, 4 keypoints)", "2 animated explainer videos (3 min each)", "Abstract concepts need visual anchoring; mobile screen is wrong medium for 400-word keypoints", "M1/lmap", "High"),
    ("Tool selection decision logic", "Reading (6 min, 5 tool descriptions)", "Interactive decision tree + reference card grid", "Comparison and selection are interactive by nature; linear reading cannot hold parallel tool attributes in working memory", "M2/l1", "High"),
    ("RICE prompt engineering framework", "Reading (8 min, 4 components)", "RICE Builder interactive (4-field input + live preview)", "Construction framework must be taught through construction, not observation", "M3/l2", "Medium"),
    ("Capabilities vs. limitations — when to trust AI", "Reading (5 min, declarative list)", "Trust Meter 5-card scenario interaction", "Trust calibration is a judgment skill; a list of rules does not train judgment", "M4/lcap", "High"),
    ("NotebookLM tool introduction", "No content (exercise only)", "Screen-capture walkthrough video (4 min)", "Employees cannot engage meaningfully with an unfamiliar tool without conceptual framing", "M5", "High"),
    ("Napkin tool introduction", "No content (exercise only)", "Screen-capture demo (90 sec)", "This tool explains itself through demonstration better than any text description", "M5", "High"),
    ("Course overview and journey map", "None", "60-second animated course map video", "Learners need orientation before starting; text course maps are rarely read", "M0", "Medium"),
    ("RICE framework — audio reinforcement", "None", "NotebookLM Audio Overview (5 min)", "Highest-value module deserves a commute-friendly audio version; low effort to produce", "M3", "Medium"),
    ("M1 foundational concepts — audio option", "None", "NotebookLM Audio Overview (5 min)", "Mobile learners who prefer audio can engage with foundational content during commute", "M1", "Medium"),
    ("M5 content — audio option (meta-use of the tool)", "None", "NotebookLM Audio Overview (3 min)", "Module about NotebookLM using NotebookLM to introduce itself is pedagogically elegant", "M5", "Medium"),
    ("Peer application stories", "None", "3 × 90-sec employee testimonial videos (phone-recorded)", "Social proof at the synthesis moment is the highest-leverage motivational intervention; peer examples model transfer", "M8", "Medium"),
    ("RICE Quick Reference Card", "Part of reading", "Keep as scannable text + downloadable PDF", "Job aid — must be findable and scannable at the desk, not watchable", "M3 (permanent)", "High"),
    ("Tool Reference Card", "Part of reading", "Keep as interactive card grid (tappable, permanent job aid)", "Reference material requires parallel scanning, not linear reading or watching", "M2 (permanent)", "High"),
    ("Data security rules and policies", "Part of reading / sim", "Keep as explicit text (required for compliance)", "Learners must be able to read, re-read, and point to specific rules", "M6 (permanent)", "High"),
    ("Post-quiz feedback and explanations", "Text (current)", "Keep as text", "Must appear immediately after an answer and be processed alongside the question context", "All modules", "High"),
]


# Sheet 3 — Assessment Improvements (item by item)
ASSESSMENT_ITEMS = [
    # (Module, Question, Current Issue, Recommended Change, Type of Change, Priority)
    ("M1", "Q2 — T/F: 'LLMs understand meaning like humans'",
     "Binary format hides comprehension level. Learner who understands 'no understanding at all' vs 'understands differently' gives same answer.",
     "Replace with scenario MCQ: 'A colleague says Copilot understood my report exactly as I would. Which is most accurate?' Test nuanced understanding of the probabilistic mechanism.",
     "Format change: T/F → MCQ", "Medium"),

    ("M1", "Q3 — Fill-blank: tokenisation (drag 'טוקנים / טוקן')",
     "4-word bank with 2 obviously wrong options (pixels, bytes). Effectively a 2-option recognition task.",
     "Replace with free-text production: 'A colleague says the AI read and understood my document. Explain in one sentence why this is misleading.' Show model answer for self-evaluation.",
     "Format change: fill-blank → free text", "High"),

    ("M1", "Q4 — Why do ChatGPT and Copilot give different answers?",
     "All 3 wrong options are immediately dismissible by anyone with basic AI experience. Zero discriminating power.",
     "Rewrite all 3 distractors with plausible misconceptions: (a) real-time internet access, (b) same knowledge base different style, (c) different training data sources with plausible-sounding specifics.",
     "Distractor rewrite", "High"),

    ("M1", "Q6 — Broadest AI category (LLM / ML / AI / ChatGPT)",
     "'ChatGPT' as a category distractor is too obviously wrong. Effectively a 3-option question.",
     "Rewrite to test hierarchical reasoning: 'A colleague says every AI mechanism is a large language model. What is wrong with this claim?' Test understanding of the hierarchy.",
     "Item rewrite (concept unchanged)", "Medium"),

    ("M2", "Q2 — Drafting task → Copilot",
     "Scenario implies generic data. Real decision depends on whether the letter contains sensitive client information.",
     "Add 'the letter includes the client's full name and account number.' Correct answer becomes JARVIS. Acknowledge that anonymise-then-Copilot is also valid in explanation.",
     "Scenario update (data sensitivity added)", "High"),

    ("M2", "Q4 — Matching: JARVIS description",
     "JARVIS matching description reflects only the Compliance Agent, not the platform.",
     "Update to: 'Internal LLM platform — all Copilot capabilities, data stays in bank. Hosts agents like the Compliance Agent.'",
     "Content accuracy update", "High"),

    ("M2", "Q5 — Fill-blank: JARVIS architecture",
     "5-word bank with 2 obviously wrong options (Copilot, POINT). Recognition task.",
     "Replace with scenario: employee asks JARVIS about current interest rates, gets a 2023 policy answer. What should they know and do? Tests RAG limitation understanding.",
     "Format change: fill-blank → scenario MCQ", "Medium"),

    ("M3", "Missing scenario item",
     "All quiz items test recall of RICE labels. No item tests whether learner can evaluate or apply the framework.",
     "Add: 'Employee sent Copilot a vague request and got a generic result. Which RICE element is most clearly missing?' Four options with plausible distractors.",
     "New item addition", "High"),

    ("M4", "Missing trust calibration item",
     "Module teaches when to trust AI output. No quiz item tests whether learner can apply this judgment.",
     "Add: 'Copilot confidently quotes a current interest rate. You're about to include it in a client document. What do you do?' Include 'ask Copilot to confirm' as a plausible wrong option.",
     "New item addition", "High"),

    ("M5", "All items — orphaned assessment",
     "With no prior instruction, the quiz tests only what learners discovered in exercises. Cannot assess comprehension of missing conceptual content.",
     "After implementing R-23 (explainer videos), rewrite to include a NotebookLM selection scenario with plausible distractors for other tools.",
     "Full rewrite (after content addition)", "Medium"),

    ("M7", "Missing application item",
     "Quiz tests recall of what Custom Instructions are, not whether learner can construct them.",
     "Add open-response item: employee's requirements described; learner writes the Custom Instruction draft. Model answer shown for self-assessment.",
     "New open-response item", "Medium"),
]


# Sheet 4 — Implementation Roadmap
ROADMAP = [
    # Sprint, Task ID, Task, Module(s), Effort, Dependencies, Type
    # ── SPRINT 0: CERTIFICATION DESIGN & LAUNCH PREP (Weeks -4 to 0) ─────────
    ("Sprint 0 (Weeks −4 to 0)", "S0-01", "Define certification passing standards (post-test threshold, mandatory modules per track, attempt limits)", "All", "Low", "L&D + HR alignment", "Policy"),
    ("Sprint 0 (Weeks −4 to 0)", "S0-02", "Design digital badge (Open Badges standard): metadata, skills claimed, expiry, verification link", "All", "Low", "Badge platform selection", "Design"),
    ("Sprint 0 (Weeks −4 to 0)", "S0-03", "Establish recertification cycle policy: 18-month expiry, trigger-based update mechanism, recertification path", "All", "Low", "L&D + IT + Legal", "Policy"),
    ("Sprint 0 (Weeks −4 to 0)", "S0-04", "Plan LMS–HR system integration for competency record and onboarding requirement", "All", "High", "IT + HR systems team", "Technical Planning"),
    ("Sprint 0 (Weeks −4 to 0)", "S0-05", "Develop manager preview brief (2-pager + 30-min team meeting agenda + WIIFM messages for 4 employee segments)", "Pre-launch", "Low", "None", "Content"),
    ("Sprint 0 (Weeks −4 to 0)", "S0-06", "Configure manager team dashboard in LMS (completion rate, module progress, post-test scores, at-risk flags, weekly email)", "LMS", "Medium", "LMS admin access", "Technical"),
    ("Sprint 0 (Weeks −4 to 0)", "S0-07", "Develop post-certification activation guide for managers (3 questions + 3 follow-up tasks + AI safety response guidance)", "Post-course", "Low", "None", "Content"),
    ("Sprint 0 (Weeks −4 to 0)", "S0-08", "Write executive sponsorship video script and produce video (3–5 min, named senior leader)", "M0", "Medium", "Executive availability", "Video Production"),
    ("Sprint 0 (Weeks −4 to 0)", "S0-09", "Design and execute 4-week communication cascade (manager brief, all-employee announcement, FAQ, role-tailored WIIFM)", "Pre-launch", "Medium", "Internal comms team", "Communication"),
    ("Sprint 0 (Weeks −4 to 0)", "S0-10", "Define and communicate completion mandate (8-week deadline, 30-day onboarding window, escalation path, support before consequences)", "All", "Low", "HR + L&D + Department heads", "Policy"),
    ("Sprint 0 (Weeks −4 to 0)", "S0-11", "Assign named L&D content owner and establish governance calendar (quarterly tool review, annual full review, trigger policy)", "All", "Low", "L&D team", "Governance"),

    # ── SPRINT 1 ──────────────────────────────────────────────────────────────
    ("Sprint 1 (Weeks 1–2)", "S1-01", "Redesign M0 diagnostic (3 dimensions, 10 items, 3 tracks)", "M0", "Medium", "None", "Content + Dev"),
    ("Sprint 1 (Weeks 1–2)", "S1-02", "Rewrite M1 Q2, Q3, Q4, Q6 (assessment fixes)", "M1", "Low", "None", "Content"),
    ("Sprint 1 (Weeks 1–2)", "S1-03", "Update M2 JARVIS description (platform vs. agent)", "M2", "Low", "None", "Content"),
    ("Sprint 1 (Weeks 1–2)", "S1-04", "Rewrite M2 Q2, Q4, Q5 (data sensitivity + JARVIS)", "M2", "Low", "S1-03", "Content"),
    ("Sprint 1 (Weeks 1–2)", "S1-05", "Add M3 RICE application scenario quiz item", "M3", "Low", "None", "Content"),
    ("Sprint 1 (Weeks 1–2)", "S1-06", "Add M3 exercise reflection field improvements", "M3", "Low", "None", "Content"),
    ("Sprint 1 (Weeks 1–2)", "S1-07", "Add bridge question after M3/l3 video", "M3", "Low", "None", "Content"),
    ("Sprint 1 (Weeks 1–2)", "S1-08", "Add M4 trust calibration scenario quiz item", "M4", "Low", "None", "Content"),
    ("Sprint 1 (Weeks 1–2)", "S1-09", "Connect M4 Copilot exercise to employee's existing use", "M4", "Low", "None", "Content"),
    ("Sprint 1 (Weeks 1–2)", "S1-10", "Add 'When Not to Use AI' section to M4", "M4", "Low", "None", "Content"),
    ("Sprint 1 (Weeks 1–2)", "S1-11", "Add M5 framing section (why use these vs. Copilot Chat)", "M5", "Low", "None", "Content"),
    ("Sprint 1 (Weeks 1–2)", "S1-12", "Update M6 EthicsCyber sim consequence text (policy references)", "M6", "Low", "Info Security team review", "Content"),
    ("Sprint 1 (Weeks 1–2)", "S1-13", "Rewrite M7 quiz to add Custom Instructions application item", "M7", "Low", "None", "Content"),
    ("Sprint 1 (Weeks 1–2)", "S1-14", "Rewrite intro video script for M0 (Copilot-first framing)", "M0", "Low", "None", "Script"),
    ("Sprint 2 (Weeks 3–6)", "S2-01", "Build M0 animated course map (60 sec)", "M0", "Low", "S1-14", "Dev + Production"),
    ("Sprint 2 (Weeks 3–6)", "S2-02", "Build 'Spot the Problem' trust calibration (5 cards)", "M1", "Medium", "None", "Dev"),
    ("Sprint 2 (Weeks 3–6)", "S2-03", "Build 'Fact-Check This' hallucination detection (5 items)", "M1", "Medium", "None", "Dev"),
    ("Sprint 2 (Weeks 3–6)", "S2-04", "Build M2 Tool Decision Flow interactive", "M2", "Medium", "S1-03", "Dev"),
    ("Sprint 2 (Weeks 3–6)", "S2-05", "Build M2 Tool Reference Card grid (permanent job aid)", "M2", "Medium", "S1-03", "Dev"),
    ("Sprint 2 (Weeks 3–6)", "S2-06", "Build RICE Builder interactive (M3)", "M3", "Medium", "None", "Dev"),
    ("Sprint 2 (Weeks 3–6)", "S2-07", "Build Trust Meter 5-card interaction (M4/lcap)", "M4", "Medium", "None", "Dev"),
    ("Sprint 2 (Weeks 3–6)", "S2-08", "Build 'What Went Wrong?' prompt diagnosis interaction (M4)", "M4", "Medium", "None", "Dev"),
    ("Sprint 2 (Weeks 3–6)", "S2-09", "Build Workflow Mapping exercise (M4)", "M4", "Medium", "None", "Dev"),
    ("Sprint 2 (Weeks 3–6)", "S2-10", "Build Data Classifier micro-tool (M6)", "M6", "Medium", "None", "Dev"),
    ("Sprint 2 (Weeks 3–6)", "S2-11", "Build Custom Instructions Builder (M7)", "M7", "Medium", "None", "Dev"),
    ("Sprint 2 (Weeks 3–6)", "S2-12", "Build Bino fallback self-guided checklist (M8)", "M8", "Medium", "None", "Dev + Content"),
    ("Sprint 2 (Weeks 3–6)", "S2-13", "Build M8 share-with-manager mechanism (Teams/email)", "M8", "Medium", "IT/Teams integration", "Dev"),
    ("Sprint 3 (Weeks 5–10)", "S3-01", "Produce M1 Video A: The AI Map animation (3 min)", "M1", "Medium", "Script ready", "Video Production"),
    ("Sprint 3 (Weeks 5–10)", "S3-02", "Produce M1 Video B: Tokens and Context animation (3 min)", "M1", "Medium", "Script ready", "Video Production"),
    ("Sprint 3 (Weeks 5–10)", "S3-03", "Generate M1 Audio Overview (NotebookLM, 5 min)", "M1", "Low", "M1 transcript", "Audio (NotebookLM)"),
    ("Sprint 3 (Weeks 5–10)", "S3-04", "Generate M3 RICE Audio Overview (NotebookLM, 5 min)", "M3", "Low", "M3 transcript", "Audio (NotebookLM)"),
    ("Sprint 3 (Weeks 5–10)", "S3-05", "Produce M5 NotebookLM screen-capture walkthrough (4 min)", "M5", "Low", "None", "Video (screen capture)"),
    ("Sprint 3 (Weeks 5–10)", "S3-06", "Produce M5 Napkin screen-capture demo (90 sec)", "M5", "Low", "None", "Video (screen capture)"),
    ("Sprint 3 (Weeks 5–10)", "S3-07", "Generate M5 Audio Overview (NotebookLM, 3 min)", "M5", "Low", "S3-05 script", "Audio (NotebookLM)"),
    ("Sprint 3 (Weeks 5–10)", "S3-08", "Add M6 video segment on remediation (60 sec)", "M6", "Low", "Script ready", "Video (add segment)"),
    ("Sprint 3 (Weeks 5–10)", "S3-09", "Produce M7 JARVIS agents context video or reading (2 min)", "M7", "Low", "None", "Video or Content"),
    ("Sprint 3 (Weeks 5–10)", "S3-10", "Film M8 peer story videos (3 × 90 sec, phone recording)", "M8", "Low", "Employee volunteers", "Video Production"),
    ("Sprint 4 (Weeks 8–12)", "S4-01", "Write and configure matched post-test (10 items)", "All", "Medium", "S1-01 diagnostic", "Content + LMS Config"),
    ("Sprint 4 (Weeks 8–12)", "S4-02", "Configure D+7 spaced retrieval quiz (M3 + M6, 6 items)", "M3, M6", "Medium", "LMS automation", "Content + LMS Config"),
    ("Sprint 4 (Weeks 8–12)", "S4-03", "Configure D+30 spaced retrieval quiz (M3 + M6, 6 items)", "M3, M6", "Medium", "LMS automation", "Content + LMS Config"),
    ("Sprint 4 (Weeks 8–12)", "S4-04", "Create manager communication kit (PDF + Teams template)", "Post-course", "Low", "None", "Content"),
    ("Sprint 4 (Weeks 8–12)", "S4-05", "Add chapter navigation to all videos over 5 min (WebVTT)", "M3, M4, M5, M6, M7", "Medium", "All videos final", "Technical"),
    ("Sprint 4 (Weeks 8–12)", "S4-06", "Generate Hebrew subtitles for all 7 videos + transcript accordions", "M1, M3, M4, M6, M7", "Medium", "All videos final", "Accessibility"),
    ("Sprint 4 (Weeks 8–12)", "S4-07", "Arabic translation — M6 + M1 first sprint", "M6, M1", "High", "Translator engaged", "Translation"),
    ("Sprint 4 (Weeks 8–12)", "S4-08", "Prompt Gallery job aid (15–20 RICE prompts by role)", "Post-course", "Medium", "M3 complete", "Content"),
    ("Sprint 4 (Weeks 8–12)", "S4-09", "xAPI migration assessment and budget planning", "Infrastructure", "High", "IT team", "Technical Planning"),

    # ── SPRINT 5: POST-LAUNCH ECOSYSTEM (Ongoing from Week 10) ───────────────
    ("Sprint 5 (Week 10+, ongoing)", "S5-01", "Launch AI Champions opt-in network (Teams channel + monthly digest + quarterly L&D check-in)", "Post-course", "Low", "Course launch complete", "Community"),
    ("Sprint 5 (Week 10+, ongoing)", "S5-02", "Build Living Resource Hub on intranet (Tool Reference Card, RICE Prompt Gallery, Data Classification Guide, What's New feed)", "Post-course", "Medium", "Intranet access + S4-08", "Content + Technical"),
    ("Sprint 5 (Week 10+, ongoing)", "S5-03", "Configure D+30 2-item pulse survey alongside spaced retrieval quiz (application + barriers questions)", "Post-course", "Low", "LMS automation + S4-03", "Content + LMS Config"),
    ("Sprint 5 (Week 10+, ongoing)", "S5-04", "Execute LMS–HR system integration (competency record sync, onboarding requirement flag)", "All", "High", "IT + HR systems team", "Technical"),
    ("Sprint 5 (Week 10+, ongoing)", "S5-05", "Publish M8 exit screen with Level 1 certification framing and Level 2 pathway reference", "M8", "Low", "Course complete", "Content + Dev"),
    ("Sprint 5 (Week 10+, ongoing)", "S5-06", "Draft 4-level AI competency architecture outline (Level 1 current, Level 2 advanced, Level 3 team lead, Specialist technical)", "Strategy", "Medium", "L&D + HR + IT", "Strategy"),
    ("Sprint 5 (Week 10+, ongoing)", "S5-07", "Configure recertification LMS trigger (18-month expiry notification + recertification path assignment)", "All", "Medium", "LMS admin + S0-03", "Technical"),
    ("Sprint 5 (Week 10+, ongoing)", "S5-08", "Configure digital badge issuance and Open Badges integration (LMS + HR system + optional LinkedIn)", "All", "Medium", "Badge platform + LMS admin", "Technical"),
]


# ─────────────────────────────────────────────────────────────────────────────
#  BUILD WORKBOOK
# ─────────────────────────────────────────────────────────────────────────────

wb = Workbook()
wb.remove(wb.active)  # remove default sheet

# ── SHEET 0: Overview ─────────────────────────────────────────────────────────
ws0 = wb.create_sheet("Overview")
ws0.sheet_view.rightToLeft = False

overview_rows = [
    ("AI Fundamentals Course Improvement Report", None),
    ("Bank Hapoalim Internal Academy · June 2026", None),
    (None, None),
    ("CONTENTS OF THIS WORKBOOK", None),
    ("Sheet", "Description"),
    ("Overview", "This page — report summary and key findings"),
    ("All Recommendations", "Master table of all 62 improvement recommendations with priority, effort, impact, and sprint"),
    ("Format Decisions", "Content format analysis — what to transform and what to keep as text"),
    ("Assessment Improvements", "Quiz item–by–item analysis and rewrites"),
    ("Implementation Roadmap", "Sprint-by-sprint task list including Sprint 0 (certification design) through Sprint 5 (ecosystem)"),
    (None, None),
    ("KEY FINDINGS", None),
    ("Finding", "Detail"),
    ("The course does not start from where employees are",
     "The M0 diagnostic collects data but routes all learners identically. No adaptive logic exists. Redesigning the diagnostic with 3 dimensions and 3 learning tracks is the highest-priority change."),
    ("Critical AI thinking skills are underdeveloped",
     "The course teaches what AI does and which tool to use. It does not adequately teach when to trust output, how to spot a hallucination in professional context, or how to reason about AI limitations in banking situations."),
    ("Too much conceptual content is delivered as passive reading on mobile",
     "5 lessons carry the conceptual load as long-form text with accordion keypoints. M1/lmap (4 dense keypoints, 6 min) and M2/l1 (5 tool descriptions, 6 min) are the highest-priority format transformations."),
    ("Assessments test recognition, not competence",
     "Most quiz items test whether employees can identify the right answer when presented. Few test whether they can produce, apply, or transfer knowledge to a real banking scenario."),
    ("JARVIS is described as a policy tool, not an internal LLM platform",
     "The course conflates JARVIS (the internal AI platform) with the Compliance Agent (one application built on it). This means employees will not use JARVIS for general tasks involving sensitive data — a compliance risk."),
    ("M5 has no conceptual content before tool exercises",
     "The only module that violates the Tell-Show-Do sequence used consistently elsewhere. Employees encounter two unfamiliar tools with no framing."),
    ("The certification has no minimum passing standard",
     "An employee who clicks through every screen at 2× speed receives the same credential as one who engaged deeply. A foundational competency certification requires a post-test threshold, mandatory module completion, and a capstone — not just time-on-screen."),
    ("Manager enablement is absent from the design",
     "The single highest predictor of training transfer is manager support. Without a preview brief, a team dashboard, and an activation conversation guide, managers are bystanders to the certification — and Level 3 transfer approaches zero."),
    ("No organizational rollout or change management plan",
     "A mandatory certification for every employee at the bank is a change management exercise, not just a content project. Without executive sponsorship, a communication cascade, and a defined accountability structure, completion will be low-quality and engagement will be low."),
    ("No post-certification ecosystem or governance model",
     "A certification without a community, updated job aids, and a content review cycle decays within 12 months. The course will reflect 2026 realities in a 2028 AI environment unless governance is designed now."),
    (None, None),
    ("TOTALS", None),
    ("Total recommendations", str(len(RECOMMENDATIONS))),
    ("High priority", str(sum(1 for r in RECOMMENDATIONS if r[7] == "High"))),
    ("Medium priority", str(sum(1 for r in RECOMMENDATIONS if r[7] == "Medium"))),
    ("Low priority", str(sum(1 for r in RECOMMENDATIONS if r[7] == "Low"))),
    ("Sprint 0 tasks (certification design + launch prep)", str(sum(1 for r in ROADMAP if "Sprint 0" in r[0]))),
    ("Sprint 1 tasks (quick wins)", str(sum(1 for r in ROADMAP if "Sprint 1" in r[0]))),
    ("Sprint 2 tasks (interactive builds)", str(sum(1 for r in ROADMAP if "Sprint 2" in r[0]))),
    ("Sprint 3 tasks (media production)", str(sum(1 for r in ROADMAP if "Sprint 3" in r[0]))),
    ("Sprint 4 tasks (infrastructure + accessibility)", str(sum(1 for r in ROADMAP if "Sprint 4" in r[0]))),
    ("Sprint 5 tasks (post-launch ecosystem + governance)", str(sum(1 for r in ROADMAP if "Sprint 5" in r[0]))),
]

ws0.column_dimensions["A"].width = 42
ws0.column_dimensions["B"].width = 80

for i, (a, b) in enumerate(overview_rows, 1):
    if a is None:
        continue
    ca = ws0.cell(row=i, column=1, value=a)
    cb = ws0.cell(row=i, column=2, value=b) if b is not None else None

    if i == 1:
        ca.font = Font(bold=True, size=16, color=RED_FILL, name="Calibri")
        ca.alignment = Alignment(horizontal="left", vertical="center")
    elif i == 2:
        ca.font = Font(bold=False, size=11, color="555555", name="Calibri")
    elif a in ("CONTENTS OF THIS WORKBOOK", "KEY FINDINGS", "TOTALS"):
        ca.fill = fill(GREY_HDR)
        ca.font = Font(bold=True, size=11, color=WHITE, name="Calibri")
        ca.alignment = Alignment(horizontal="left", vertical="center")
        if cb:
            cb.fill = fill(GREY_HDR)
    elif a in ("Sheet", "Finding", "Total recommendations"):
        ca.fill = fill(MID_GREY)
        ca.font = Font(bold=True, size=10, name="Calibri")
        if cb:
            cb.fill = fill(MID_GREY)
            cb.font = Font(bold=True, size=10, name="Calibri")
    else:
        ca.font = body_font(size=10)
        ca.alignment = wrap()
        if cb:
            cb.font = body_font(size=10)
            cb.alignment = wrap()
        # alternate rows
        bg = LIGHT_GREY if i % 2 == 0 else WHITE
        ca.fill = fill(bg)
        if cb:
            cb.fill = fill(bg)

ws0.row_dimensions[1].height = 30
for r in range(3, len(overview_rows)+1):
    ws0.row_dimensions[r].height = 40


# ── SHEET 1: All Recommendations ─────────────────────────────────────────────
ws1 = wb.create_sheet("All Recommendations")

COLS1 = ["ID", "Category", "Module", "Current State",
         "Issue / Gap", "Recommendation", "New Format / Asset",
         "Priority", "Effort", "Impact", "Sprint"]
WIDTHS1 = [8, 22, 14, 38, 44, 55, 32, 10, 10, 10, 18]

write_header_row(ws1, 1, COLS1, bg=RED_FILL, fg=WHITE, size=10)
set_col_widths(ws1, WIDTHS1)
ws1.row_dimensions[1].height = 30
ws1.freeze_panes = "A2"

for row_idx, rec in enumerate(RECOMMENDATIONS, 2):
    priority = rec[7]
    bg = priority_fill(priority)
    for col_idx, val in enumerate(rec, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=val)
        cell.fill = fill(bg)
        cell.font = body_font(bold=(col_idx == 1))
        cell.alignment = wrap()
        cell.border = border()
        if col_idx == 8:  # Priority column
            color_map = {"High": "B71C1C", "Medium": "E65100", "Low": "2E7D32"}
            cell.font = Font(bold=True, size=10, color=color_map.get(priority, "000000"), name="Calibri")
    ws1.row_dimensions[row_idx].height = 80


# ── SHEET 2: Format Decisions ─────────────────────────────────────────────────
ws2 = wb.create_sheet("Format Decisions")

COLS2 = ["Content", "Current Format", "Recommended Format", "Rationale", "Module", "Priority"]
WIDTHS2 = [35, 28, 38, 58, 14, 10]

write_header_row(ws2, 1, COLS2, bg=RED_FILL, fg=WHITE, size=10)
set_col_widths(ws2, WIDTHS2)
ws2.row_dimensions[1].height = 30
ws2.freeze_panes = "A2"

# Section headers
KEEP_ROWS = {10, 11, 12, 13, 14, 15, 16}  # rows that are "keep as text"

for row_idx, fd in enumerate(FORMAT_DECISIONS, 2):
    priority = fd[5]
    is_keep = "Keep" in fd[2]
    bg = "#EAF4E9" if is_keep else priority_fill(priority)
    for col_idx, val in enumerate(fd, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=val)
        cell.fill = fill(bg[1:] if bg.startswith("#") else bg)
        cell.font = body_font()
        cell.alignment = wrap()
        cell.border = border()
    ws2.row_dimensions[row_idx].height = 65

# Add section label rows
ws2.insert_rows(2)
for c in range(1, 7):
    cell = ws2.cell(row=2, column=c, value="TRANSFORM — Replace with higher-impact format" if c == 1 else "")
    cell.fill = fill("B71C1C")
    cell.font = Font(bold=True, size=10, color=WHITE, name="Calibri")
    cell.border = border()

ws2.insert_rows(13)
for c in range(1, 7):
    cell = ws2.cell(row=13, column=c, value="KEEP — Text is the right format here" if c == 1 else "")
    cell.fill = fill("2E7D32")
    cell.font = Font(bold=True, size=10, color=WHITE, name="Calibri")
    cell.border = border()


# ── SHEET 3: Assessment Improvements ─────────────────────────────────────────
ws3 = wb.create_sheet("Assessment Improvements")

COLS3 = ["Module", "Question / Item", "Current Issue", "Recommended Change", "Type of Change", "Priority"]
WIDTHS3 = [10, 38, 50, 58, 28, 10]

write_header_row(ws3, 1, COLS3, bg=RED_FILL, fg=WHITE, size=10)
set_col_widths(ws3, WIDTHS3)
ws3.row_dimensions[1].height = 30
ws3.freeze_panes = "A2"

for row_idx, item in enumerate(ASSESSMENT_ITEMS, 2):
    priority = item[5]
    bg = priority_fill(priority)
    for col_idx, val in enumerate(item, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=val)
        cell.fill = fill(bg)
        cell.font = body_font(bold=(col_idx in (1, 5)))
        cell.alignment = wrap()
        cell.border = border()
    ws3.row_dimensions[row_idx].height = 90


# ── SHEET 4: Implementation Roadmap ──────────────────────────────────────────
ws4 = wb.create_sheet("Implementation Roadmap")

COLS4 = ["Sprint", "Task ID", "Task Description", "Module(s)", "Effort", "Dependencies", "Type"]
WIDTHS4 = [26, 10, 62, 18, 10, 32, 22]

write_header_row(ws4, 1, COLS4, bg=RED_FILL, fg=WHITE, size=10)
set_col_widths(ws4, WIDTHS4)
ws4.row_dimensions[1].height = 30
ws4.freeze_panes = "A2"

sprint_colors = {
    "Sprint 0": "FCE4EC",  # pink light — pre-launch
    "Sprint 1": "FFF3E0",  # amber light
    "Sprint 2": "E3F2FD",  # blue light
    "Sprint 3": "F3E5F5",  # purple light
    "Sprint 4": "E8F5E9",  # green light
    "Sprint 5": "E0F7FA",  # teal light — post-launch
}
sprint_header_colors = {
    "Sprint 0": "880E4F",  # deep pink — pre-launch
    "Sprint 1": "E65100",
    "Sprint 2": "1565C0",
    "Sprint 3": "6A1B9A",
    "Sprint 4": "2E7D32",
    "Sprint 5": "00695C",  # teal — post-launch
}

current_sprint = None
actual_row = 2
for task in ROADMAP:
    sprint = task[0]
    # Insert sprint header if new sprint
    if sprint != current_sprint:
        current_sprint = sprint
        hcol = sprint_header_colors.get(sprint[:8], GREY_HDR)
        for c in range(1, 8):
            cell = ws4.cell(row=actual_row, column=c, value=sprint if c == 1 else "")
            cell.fill = fill(hcol)
            cell.font = Font(bold=True, size=11, color=WHITE, name="Calibri")
            cell.border = border()
        ws4.row_dimensions[actual_row].height = 22
        actual_row += 1

    bg = sprint_colors.get(sprint[:8], LIGHT_GREY)
    for col_idx, val in enumerate(task, 1):
        cell = ws4.cell(row=actual_row, column=col_idx, value=val)
        cell.fill = fill(bg)
        cell.font = body_font(bold=(col_idx == 2))
        cell.alignment = wrap()
        cell.border = border()
    ws4.row_dimensions[actual_row].height = 40
    actual_row += 1


# ─────────────────────────────────────────────────────────────────────────────
#  SAVE
# ─────────────────────────────────────────────────────────────────────────────
out_path = "/Users/avilevi/Documents/projects/AI Fundumentals/ai-fundamentals/docs/course-improvement-report.xlsx"
wb.save(out_path)
print(f"✓ Saved: {out_path}")
print(f"  Sheets: {[ws.title for ws in wb.worksheets]}")
print(f"  Recommendations: {len(RECOMMENDATIONS)}")
print(f"  Roadmap tasks:   {len(ROADMAP)}")
